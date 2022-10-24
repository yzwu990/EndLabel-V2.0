# -*- coding: utf-8 -*-
"""
Created on Mon Jan 24 12:44:02 2022

@author: Yz Wu
"""
"""
Version V2.0 updated on Mon Jan 24 2022

Support new Booho End Label format

"""
"""
Version V1.1 updated on Wed Nov 03 2021

Fix: Can't creat endlabels when packing list has less than 3 "序号"

"""





##导包

import numpy as np 
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

import tkinter as tk 
from tkinter import filedialog
import tkinter.messagebox
from tkinter import *

#建立窗口window
window = tk.Tk()
 
#给窗口的可视化起名字
window.title('箱贴生成程序 V2.0')
 
#设定窗口的大小(长 x 宽)
window.geometry('500x280')  
#设定程序图标
window.iconbitmap('maxonn.ico')



## 设定标签
#装箱单标签
var_packingList = tk.StringVar(value="1.点击按钮导入装箱单")    # 将label标签的内容设置为字符类型，用var来接收此函数的传出内容用以显示在标签上
l_packingList = tk.Label(window, textvariable=var_packingList, bg='green', fg='white', font=('Arial', 12), width=20, height=2)
# 说明： bg为背景，fg为字体颜色，font为字体，width为长，height为高，这里的长和高是字符的长和高，比如height=2,就是标签有2个字符高
l_packingList.place(x=30, y=20)

#箱贴标签
var_template = tk.StringVar(value="2.点击按钮导入箱贴模板")  
l_template = tk.Label(window, textvariable=var_template, bg='red', fg='white', font=('Arial', 12), width=20, height=2)
l_template.place(x=270, y=20)

#Supplier Booking Ref 标签
var_booking = tk.StringVar(value="3.输入订舱号后点击确定")  
l_booking = tk.Label(window, textvariable=var_booking, bg='blue', fg='white', font=('Arial', 12), width=20, height=2)
l_booking.place(x=30, y=140)

#生成箱贴
var_generate = tk.StringVar(value="4.生成箱贴")  
l_generate = tk.Label(window, textvariable=var_generate, bg='black', fg='white', font=('Arial', 12), width=20, height=2)
l_generate.place(x=270, y=140)


##定义Function，既点击按钮时执行的程序
#导入装箱单模板按钮
def hit_me_p():

    var_packingList.set('装箱单已导入')
    #全局变量，获取装箱单路径
    global packingList
    packingList = filedialog.askopenfilename()

#导入箱贴模板按钮
def hit_me_t():

    var_template.set('箱贴模板已导入')
    #全局变量，获取箱贴模板路径
    global templateBook
    templateBook = filedialog.askopenfilename()
    

#生成箱贴按钮
def generate():
    
    #导入装箱单
    packingList1 = load_workbook(packingList)
    #定位装箱单中的工作表
    sheet = packingList1.worksheets[0]
    
    #导入模板
    templateBook1 = load_workbook(templateBook)
    #定位模板文件中的工作表
    templateSheet = templateBook1.worksheets[0]   
    
    cells = sheet['G7:''G'+str(sheet.max_row)]
    #读取箱号列
    values = [[cell.value for cell in columns] for columns in cells]
    #降维
    b = np.squeeze(values)

    #定位空值
    p=np.where(b==None)


 
    #空值个数
    Blank=len(p[0])
    
    #准备一个空集，放n组装箱单
    P=[]
##########       
    # #产生第一组箱号
    # p_0=b[0:(p[0][0])]
    
    # #把第一组箱号存到P中
    # P+=[p_0]
    
    
    # #产生第2到n-1组箱号
    # for i in range(1,Blank) :
    #     names='p_'+str(i)
    #     locals()['p_'+str(i)]=b[p[0][i-1]+1:p[0][i]]
    #     #把产生的箱号组存到P中
    #     P+=[b[p[0][i-1]+1:p[0][i]]]
    
    # #产生第n组箱号
    # p_l=b[p[0][i]+1:]
    
    # #把第n组箱号存到P中
    # P+=[p_l]
    
    # #print(P[3][0])
############################
    if Blank == 0:
        # print('b[0:(p[0][0])]=',b[0:(p[0][0])])
        values+=[[None]]
        b = np.squeeze(values)
        p=np.where(b==None)
        # print('b=',b)
        # print('p=',p)
        p_0=b[0:(p[0][0])]
        P+=[p_0]
    elif Blank == 1:
        p_0=b[0:(p[0][0])]
        P+=[p_0]
        p_l=b[p[0][0]+1:]
        P+=[p_l]
        # print('p_l=',p_l)
    else:
      
        #产生第一组箱号
        p_0=b[0:(p[0][0])]
        
        # print('p_0=',p_0)
        
        #把第一组箱号存到P中
        P+=[p_0]
        # print('P=',P)
        
        #产生第2到n-1组箱号
        for i in range(1,Blank) :
            names='p_'+str(i)
            locals()['p_'+str(i)]=b[p[0][i-1]+1:p[0][i]]
            #把产生的箱号组存到P中
            P+=[b[p[0][i-1]+1:p[0][i]]]
        
        #产生第n组箱号
        p_l=b[p[0][i]+1:]
        
        #把第n组箱号存到P中
        P+=[p_l]
        
        # #print(P[3][0])
    
    # print('P=',P)








    
    ##以上，P[n]即为第n组箱号（n=0,1,...n）##
    sequence = len(P)+1
    
    #所有合并的单元格
    merged_ranges = sheet.merged_cell_ranges
    
    #输入箱号的范围
    for n in range(0,len(P)+1):
        # 创建一个工作簿对象
        wb = Workbook()
        #n的取值要小于数组长度，否则报错
        if n < len(P):     
            #循环创建并命名工作表
            if len(P[n]) !=1:      
                for i in range(P[n][0],P[n][len(P[n])-1]+1) :
                    wb.create_sheet(title='箱号'+str(i))
                        

                    Number="箱号"+str(i)
                    ws=wb[Number]
                    # for image in templateSheet._images:
                    #     ws.add_image(image)
                    #     WorkbookName='箱号'+str(P[n][0])+'-'+str(P[n][len(P[n])-1])+'.xlsx'
                    #     wb.save(WorkbookName)
                    #     wb.close()
                    
                    
                    
                    
                    ##从模板复制
                    from copy import copy
                
                    for row in templateSheet.rows:
                        for cell in row:
                            new_cell = ws.cell(row=cell.row, column=cell.col_idx,
                                    value= cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                    ##
                    ## 通过箱号定位，填写当前工作表 
                    for cellss in cells:
                        for cell in cellss:
                            if cell.value ==i:
                                #箱号i所对应的PO
                                coor_B='B'+str(cell.row)
                                #箱号i所对应的款号
                                coor_C='C'+str(cell.row)
                                #箱号i所对应的颜色_中文
                                coor_K='K'+str(cell.row)
                                #箱号i所对应的颜色_英文
                                coor_L='L'+str(cell.row)
                                
                            
                                #循环判断PO是否为合并单元格，并把PO号填入当前工作表
                                for merged_range_B in merged_ranges:
                                    if sheet[coor_B].coordinate in merged_range_B:
                                        merged_value = merged_range_B.start_cell.value
                                        ws['C4']=merged_value
                                        break
                                    else:
                                        PO=sheet['B'+str(cell.row)]
                                        ws['C4']=PO.value
                                        
                                #循环判断款号是否为合并单元格，并把款号填入当前工作表
                                for merged_range_C in merged_ranges:
                                    if sheet[coor_C].coordinate in merged_range_C:
                                        merged_value = merged_range_C.start_cell.value
                                        ws['C6']=merged_value
                                        break
                                    else:
                                        Style=sheet['C'+str(cell.row)]
                                        ws['C6']=Style.value
                              
                                #循环英文颜色是否为合并单元格，并把英文颜色填入当前工作表
                                for merged_range_L in merged_ranges:
                                    if sheet[coor_L].coordinate in merged_range_L:
                                        merged_value = merged_range_L.start_cell.value
                                        ws['C8']=merged_value
                                        break
                                    else:
                                        Colour_en=sheet['L'+str(cell.row)]
                                        ws['C8']=Colour_en.value               

                                ##填写其他信息
                                Size=sheet['E'+str(cell.row)]
                                Quantity=sheet['F'+str(cell.row)]
                           
                                ws['C7']=Size.value
                            
                                ws['C9']=Quantity.value
                                if b[-1]!= None:
                                    ws['C10']=str(i) + " of " + str(b[-1])
                                else: 
                                    ws['C10']=str(i) + " of " + str(len(b)-1)
                                    
                                #订舱号    
                                ws['C5']=booking_value    
                                    
                                    
                                    
                                ##
                                # #调整D，F，E列的宽度，达到和模板一样
                                ws.column_dimensions['C'].width = 25
                                # ws.column_dimensions['F'].width = 5.5
                                # ws.column_dimensions['E'].width = 8.5
                                # ws.row_dimensions[16].height = 25.7
                                
            
                                
                                ##
                                #调整居中
                                ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
                                ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
                                ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
                                ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
                                ws['C9'].alignment = Alignment(horizontal='center', vertical='center')
                                ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
                                ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
                            


                            
                            
                            
                          
                #删除多余的"Sheet"页面
                del wb["Sheet"]      
                #工作簿名称
                WorkbookName='箱号'+str(P[n][0])+'-'+str(P[n][len(P[n])-1])+'.xlsx'
                #保存工作薄
                wb.save(WorkbookName)
                # 最后关闭文件
                wb.close()
               
            else:
                print(n)
                #循环创建并命名工作表
                wb.create_sheet(title='箱号'+str(P[n][0]))
                Number="箱号"+str(P[n][0])
                ws=wb[Number]
                             
                ##从模板复制
                from copy import copy
            
                for row in templateSheet.rows:
                    for cell in row:
                        new_cell = ws.cell(row=cell.row, column=cell.col_idx,
                                value= cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                ##
                ## 通过箱号定位，填写当前工作表        
                for cellss in cells:
                    for cell in cellss:
                        if cell.value ==P[n][0]:
                            
               
                            PO=sheet['B'+str(cell.row)]
                            Style=sheet['C'+str(cell.row)]
                            Size=sheet['E'+str(cell.row)]
                            Quantity=sheet['F'+str(cell.row)]
                            Colour_zh=sheet['K'+str(cell.row)]
                            Colour_en=sheet['L'+str(cell.row)]
                    
                            ws['C4']=PO.value
                            ws['C6']=Style.value
                            ws['C7']=Size.value
                            ws['C8']=Colour_en.value
                            ws['C9']=Quantity.value
                            ws['C10']=str(P[0][n]) + " of " + str(b[-1])
                            #订舱号    
                            ws['C5']=booking_value    
                                    
                            
                            ##
                            #调整D，F，E列的宽度，达到和模板一样
                            ws.column_dimensions['D'].width = 20.5
                            ws.column_dimensions['F'].width = 5.5
                            ws.column_dimensions['E'].width = 8.5
                            ws.row_dimensions[16].height = 25.7
                            
                            ##
                            #调整居中
                            ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
                            ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
                            ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
                            ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
                            ws['C9'].alignment = Alignment(horizontal='center', vertical='center')
                            ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
                            ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
                            
                ##
                #删除多余的"Sheet"页面
                # del wb["Sheet"]                     
                # #工作簿名称
                # WorkbookName='箱号'+str(P[0][n])+'.xlsx'
                # #保存工作薄
                # wb.save(WorkbookName)
                # 最后关闭文件
                wb.close()
    # print('')
    # print('**************')
    # print('箱贴生成完毕')
    tkinter.messagebox.showinfo(title='完成', message='恭喜您！箱贴生成完毕！')


## 放置Button
#导入装箱单按钮
b_packingList = tk.Button(window, text='1.导入装箱单', font=('Arial', 12), width=10, height=1, command=hit_me_p)
b_packingList.place(x=70, y=80)
#导入箱贴按钮
b_template = tk.Button(window, text='2.导入箱贴模板', font=('Arial', 12), width=12, height=1, command=hit_me_t)
b_template.place(x=300, y=80)

#输入Supplier Booking Ref
enty = tk.Entry(window, width=18,relief=GROOVE,bd=5)
enty.place(x=50, y=200)


def booking():
    global booking_value
    booking_value = enty.get()
    var_booking.set('订舱号为：'+ booking_value)
    
b_booking = tk.Button(window,text='3.确定',font=('Arial', 12), width=10, height=1,command=booking)
b_booking.place(x=65, y=240)

#生成箱贴
b_ge = tk.Button(window, text='4.生成箱贴', font=('Arial', 12), width=12, height=1, command=generate)
b_ge.place(x=300, y=210)

 


#主窗口循环显示
window.mainloop()

